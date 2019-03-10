'''
Project Name: InventoryEXT_SMV_GUI - Inventory To Simplified Model View, Code Material Design
Proper Name: Excel Data Inventory To Simplified Model ListView
Version: unknown.beta
Authors: Janrey Tuazon Licas - App Design (Semi)
        Sam Matienzo (Database Worker)
GUI Library Used: Kivy
    Base Theme: KivyMD (Material Design Preset Designs)

Date Initialization: 02/21/2019
Date Finished (First Time): 03/10/2019 @ 11:00 PM / 23:00

'''
from kivy.app import App
from kivy.lang import Builder
from kivy.metrics import dp
from kivy.core.window import Window
from kivy.uix.image import Image
from kivy.uix.widget import Widget
from kivy.uix.screenmanager import ScreenManager, Screen, FadeTransition
from kivy.config import Config
from kivy.logger import Logger as KivyDebugger_Print
from kivymd.button import MDIconButton
from kivymd.list import ILeftBody, ILeftBodyTouch, IRightBodyTouch, BaseListItem
from kivymd.selectioncontrols import MDCheckbox
from kivymd.snackbar import Snackbar
from kivymd.theming import ThemeManager
from kivy.utils import get_color_from_hex
from kivymd.color_definitions import colors
from kivy.clock import Clock
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from random import choice as RandomizePick
from os.path import isfile as FileExist
import sqlite3

class MD_InventoryEXDI_SMLV_GUI(App):
    """
        Variables that plays the main role for changing properites. Some holds the path
        and some holds specific data to be sent on initialization
    """
    theme_cls = ThemeManager()
    DataIdentity = [Element for Element in range(4)] # For Storing User Identity
    title = "Excel Data Inventory To Simplified Model ListView | Data Editor" #Excel to Simplified Model VIew in GUI
    Icon_Status = 'brightness-1' # Used on initializing toolbar with variable to be able to change the icon on_press of the function
    SQLite_MainPath = 'EXI_SMV_AppData.db' #Main FileName of SQL DB
    Active_DataElement = '' # Active DataElement is used to bind IDS of ListItemelement. It was also used to return values when applyind changes, see ExcelLoadDataTextField_OnClick for structure of the code.
    """
        This DataReport Variables is placed outside any functions, because this will be used by some functions
        that shouldn't be override values for each data bind.
    """
    DataReport_TotalUnits = 0
    DataReport_CalculatedProposed = 0
    DataReport_ActiveOnHand = 0
    DataReport_AverageOnHand = 0
    DataReport_AverageQuarterAllotment_1 = 0
    DataReport_AverageQuarterAllotment_2 = 0
    DataReport_AverageQuarterAllotment_3 = 0
    DataReport_AverageQuarterAllotment_4 = 0
    DataReport_AverageQuarterAllotment = 0
    DataReport_Calculated_TotalCost = 0
    DataReport_Calculated_UnitCost = 0
    """
    (System Only Method, Function) => def build(self), The Application Builder that uses KV File
        This function initializes the stylesheet (that's what we call) or let's say the KV File that contains the whole look of the UI
        We use Builder from Kivy to load the KV file as a string... Basically we can load in three methods,
        First Method: Builder.load_string(variable), where variable contains the whole structure of the UI
        Second Method: Builder.load_file(fileName), where file contains the whole structure of the UI
        Third Method: We dont know, but all we know is that we dont use Builder to laod the UI, but rather using class method with return
    """
    def build(self):
    # rand_primary_palette is a copy of theme_cls.primary_palette that has Option Property. It was copied due to random.RandomizePick only produce '1'.
        # NOTE: We dont know the color preference of the client so we randomize it.
        rand_primary_palette = ['Pink', 'Blue', 'Indigo', 'BlueGrey', 'Brown','LightBlue','Purple', 'Yellow', 'DeepOrange','Green', 'Red', 'Teal', 'Orange', 'Cyan', 'Amber','DeepPurple', 'Lime']
        self.MainClassBuildFile = Builder.load_file("MD_DesignClass_File.kv")
        self.theme_cls.primary_palette = (RandomizePick(rand_primary_palette))
        self.theme_cls.primary_hue = '400'
        self.theme_cls.theme_style = 'Light'
        return self.MainClassBuildFile
    """
    (System Only Method, Function) => def on_start(self, no any other arguments passed.)
        A Single Method used on App Parent Class to handle extra initialization.
        When initializing the program, we cannot access a set of dictionaries such as self.root.ids
        Where specific data changing function such as MDNavDrawer_UserDefinition and MDScreenManagement_CheckVal uses self.root.ids
        to change properties. This was initialized so that on start things have to be on set before on use.
    """
    def on_start(self):
        Database_Iterator = 0
        Base_Iterator = 1
        try:
            SQLData_Check = sqlite3.connect(self.SQLite_MainPath)
            SQLData_Cursor = SQLData_Check.cursor()
            SQLData_Cursor.execute('SELECT * FROM user_TableLastSaveString')
            SQLData_PlaceHolder = SQLData_Cursor.fetchone()
            for EachData in range(7):
                if Base_Iterator == 1:
                    self.root.ids.DataBind_FilePath.text = str(SQLData_PlaceHolder[Database_Iterator])
                elif Base_Iterator == 2:
                    self.root.ids.DataBind_FileExportPath.text = str(SQLData_PlaceHolder[Database_Iterator])
                elif Base_Iterator == 3:
                    self.root.ids.DataBind_ColumnCellStartPoint.text = str(SQLData_PlaceHolder[Database_Iterator])
                elif Base_Iterator == 4:
                    self.root.ids.DataBind_ColumnCellEndPoint.text = str(SQLData_PlaceHolder[Database_Iterator])
                elif Base_Iterator == 5:
                    self.root.ids.DataBind_RowCellStartPoint.text = str(SQLData_PlaceHolder[Database_Iterator])
                elif Base_Iterator == 6:
                    self.root.ids.DataBind_RowCellEndPoint.text = str(SQLData_PlaceHolder[Database_Iterator])
                elif Base_Iterator == 7:
                    self.root.ids.DataBind_ExcelFileSheet.text = str(SQLData_PlaceHolder[Database_Iterator])
                Base_Iterator += 1
                Database_Iterator += 1
            SQLData_Check.close()
            self.MDNavDrawer_UserDefinition() # Sets User Identification in Navigation
            self.MDScreenManagement_CheckVal() # Checks SQL user_isFirstTime if it is True or False which returns specific screen manager current value.
        except:
            self.root.ids.DataBind_FilePath.text = ''
            self.root.ids.DataBind_FileExportPath.text = ''
            self.root.ids.DataBind_ColumnCellStartPoint.text = ''
            self.root.ids.DataBind_ColumnCellEndPoint.text = ''
            self.root.ids.DataBind_RowCellStartPoint.text = ''
            self.root.ids.DataBind_RowCellEndPoint.text = ''
            self.root.ids.DataBind_ExcelFileSheet.text = ''
            self.MDNavDrawer_UserDefinition() # Sets User Identification in Navigation
            self.MDScreenManagement_CheckVal() # Checks SQL user_isFirstTime if it is True or False which returns specific screen current value.
            SQLData_Check.close()
    
    """
    (System Only Method, Function) => def on_stop(self, no any other arguments passed.)
        A Single Method used on App Parent Class to handle extra initialization.
        This event has been handle due to SQL Saving strings, meaning making save states to avoid filling up the textfield again.
    """
    def on_stop(self):
        Iterator = 0
        SQLData_Check = sqlite3.connect(self.SQLite_MainPath)
        SQLData_Cursor = SQLData_Check.cursor()
        SQLData_Cursor.execute('UPDATE user_TableLastSaveString SET string_ExcelImportPath = ?, string_ExcelExportPath = ?, string_ExcelColumnStart = ?, string_ExcelColumnEnd = ?, string_ExcelRowStart = ?, string_ExcelRowEnd = ?, string_ExcelSheetName = ?', (self.root.ids.DataBind_FilePath.text, self.root.ids.DataBind_FileExportPath.text, self.root.ids.DataBind_ColumnCellStartPoint.text, self.root.ids.DataBind_ColumnCellEndPoint.text, self.root.ids.DataBind_RowCellStartPoint.text, self.root.ids.DataBind_RowCellEndPoint.text, self.root.ids.DataBind_ExcelFileSheet.text))
        SQLData_Check.commit()
        SQLData_Check.close()
    """
    FUNCTION PROGRAM INITIALIZERS AND CHECKPOINT SECTION6
        One of the main components of the program, it initializes the program in a very certian way. There is also some function
        that involves checkpoint that sets specific datas on SQL before proceeding.

    (Callable Method, Function) => def on_start(self) Overrides config.ini
        This was meant to run everytime, no checks intended to be made due to the fact that it can break the app
        when every changes has been made. Not sure why it happens but whatever.
    """
    def KivyConfig_Setup(self): 
        try:
            #This Config is loaded by default and must be loaded everytime.
            Config.set('graphics','width','1024')
            Config.set('graphics','height','768')
            Config.set('graphics','show_cursor','1')
            Config.set('graphics','minimum_width','1024')
            Config.set('graphics','minimum_height','768')
            Config.set('kivy','exit_on_escape','0')
            Config.write()
            #Returns String for CriticalComponent_Check Function whether to proceed or not,
            return 'Passed'
        except Exception as ErrorMessage:
            self.MDUserNotif_SnackbarHandler(str(ErrorMessage))
            return 'Failed'
    """
    (Callable Method, Function) => def SQLite_Generate_FirstTime(self), Only Checks File Exist
        Maybe returns True Or False Like Values to determine to create a file or not
        Things we can't do is to check its file size or even the content in order for it to be valid.
    """
    def SQLite_Generate_FirstTime(self):
        if FileExist(self.SQLite_MainPath):
            return 'Exist'
        else:
            return 'DoesNotExist'
        pass
    """
    (Callable Method, Function) => def CriticalComponent_Check(self), The Gate to Initialize
        Checks SQL Database for Data and Creates Data for Initialization, Sets Variables @ SQL
        NOTE: KivyDebugger_Print is part of the Kivy, it was used to print out specific data, just like print but with formatting
    """
    def CriticalComponent_Check(self):
        if self.KivyConfig_Setup() == 'Passed':
            KivyDebugger_Print.info('Kivy Configuration Load: Kivy Appplication Parameters has successfully passed parameters.')
            if self.SQLite_Generate_FirstTime() == 'Exist':
                KivyDebugger_Print.info('SQLite Data Load: File Exist, Data Loaded...')
                return 1
            else:
                KivyDebugger_Print.warning('SQLite Data Failed: Attempting to Create a Database under application directory...')
                try:
                    SQL_Connect = sqlite3.connect(self.SQLite_MainPath)
                    SQL_Connect.execute('CREATE TABLE IF NOT EXISTS user_TableDatabase(user_FirstName STRING, user_LastName STRING, user_JobCurrent STRING, user_Password STRING, user_isFirstTime )')
                    SQL_Connect.execute('CREATE TABLE IF NOT EXISTS user_TableLastSaveString(string_ExcelImportPath STRING, string_ExcelExportPath STRING, string_ExcelColumnStart STRING, string_ExcelColumnEnd STRING, string_ExcelRowStart STRING, string_ExcelRowEnd STRING, string_ExcelSheetName STRING)')
                    SQL_Connect.commit()
                    SQL_Connect.close()
                    KivyDebugger_Print.info('SQL Data Creation: Data has been succesfully created from the database!')
                    return 1
                except:
                    KivyDebugger_Print.error('SQL Database Write: SQL Failed to Execute. Application will be terminated...') 
                    return 0
        else:
            KivyDebugger_Print.warning('Kivy Configuration Load: Kivy Appplication Parameters has failed to comply its execution. Continuing....')
            if self.SQLite_Generate_FirstTime() == 'Exist':
                KivyDebugger_Print.info('SQLite Data Load: File Exist, Data Loaded...')
                return 1
            else:
                KivyDebugger_Print.warning('SQLite Data Load Failed: Attempting to Create a Database under application directory...')
                #try:
                SQL_Connect = sqlite3.connect(self.SQLite_MainPath)
                SQL_Connect.execute('CREATE TABLE IF NOT EXISTS user_TableDatabase(user_FirstName STRING, user_LastName STRING, user_JobCurrent STRING, user_Password STRING, user_isFirstTime INTEGER)')
                SQL_Connect.execute('CREATE TABLE IF NOT EXISTS user_TableLastStringSave(string_ExcelImportPath STRING, string_ExcelExportPath STRING, string_ExcelColumnStart STRING, string_ExcelColumnEnd STRING, string_ExcelRowStart STRING, string_ExcelRowEnd STRING, string_ExcelSheetName STRING, string_ShowFormulae STRING)')
                SQL_Connect.commit()
                SQL_Connect.close()
                KivyDebugger_Print.info('SQL Data Creation: Data has been succesfully created from the database!')
                return 1
    """
    (Callable Method, Function) => def MDScreenManagement_CheckVal(self), Function Aware of Screen To Choose Based on SQL Database
        The Function identifies if a specific data has any value, this applies user_isFirstTime value to avoid going back to the MainWindow_FirsTime_GetIdentity Screen...
    """
    def MDScreenManagement_CheckVal(self):
        SQLData_Check = sqlite3.connect(self.SQLite_MainPath)
        SQLData_Cursor = SQLData_Check.cursor()
        SQLData_Cursor.execute('SELECT user_isFirstTime FROM user_TableDatabase WHERE user_isFirstTime=1 or user_isFirstTime=0')
        if SQLData_Cursor.fetchone():
            self.root.ids.Edit_SectionAccess.disabled = False
            self.root.ids.DataManagement_SectionAccess.disabled = False
            self.root.ids.ScreenFrame_MainWindow.current = 'MainWindow_SimplifiedView'
        else:
            self.root.ids.ScreenFrame_MainWindow.current = 'Startup_FirstTime_GetInfo'
        SQLData_Check.close()
    """
    (Callable Method, Function) => def MDNavDrawer_UserDefinition(self), Function Aware on User's Identity
        What it can't do, awareness of NULL string of the user identity which cause to concatenate NULL,NULL name on the self.root.ids.nav_drawer.text
    """
    def MDNavDrawer_UserDefinition(self):
        Iterator = 0
        SQLData_Check = sqlite3.connect(self.SQLite_MainPath)
        SQLData_Cursor = SQLData_Check.cursor()
        SQLData_Cursor.execute('SELECT * FROM user_TableDatabase')
        SQLData_Fetch = SQLData_Cursor.fetchone()
        try:
            for EachData in range(4):
                if len(SQLData_Fetch[Iterator]) <= 1:
                    self.DataIdentity[Iterator] = 'Unknown'
                    Iterator +=1
                else:
                    self.DataIdentity[Iterator] = SQLData_Fetch[Iterator]
                    Iterator += 1
        except:
            self.DataIdentity[0] = 'Unknown'
            self.DataIdentity[1] = 'User'
            self.DataIdentity[2] = 'Unknown'
        self.root.ids.nav_drawer.header_name = 'Hello, {0} {1}!'. format(self.DataIdentity[0], self.DataIdentity[1])
        self.root.ids.nav_drawer.subheader_name = 'Position: {0}'.format(self.DataIdentity[2])
    """
    (Callable Method, Function) => def MD_FirstTime_DataSubmission(self), First Screen Initializer for First Times
        This is a setup that needs SQL Database to store the user data so that on the next launch, the data will only be loaded.
    """
    def MD_FirstTime_DataSubmission(self):
        if len(self.root.ids.FirstTimer_DataFirstName.text) <= 1 or len(self.root.ids.FirstTimer_DataLastName.text) <= 1 or len(self.root.ids.FirstTimer_DataJobPosition.text) <= 1:
            self.MDUserNotif_SnackbarHandler('Error, TextInputs should have characters, 2 or more.')
        else:
            SQLData = sqlite3.connect(self.SQLite_MainPath)
            SQLData_Cursor = SQLData.cursor()
            """
                One of these execute() uses the data on textfields, why the other one creates only for on_stop(self) function.
                This was intended so that code will be minimal on on_stop() system call.
            """
            SQLData.execute('INSERT INTO user_TableDatabase(user_FirstName, user_LastName, user_JobCurrent, user_Password, user_isFirstTime) values (?, ?, ?, ?, ?)', (self.root.ids.FirstTimer_DataFirstName.text, self.root.ids.FirstTimer_DataLastName.text, self.root.ids.FirstTimer_DataJobPosition.text,self.root.ids.FirstTimer_DataPassword.text, 0))
            SQLData.execute('INSERT INTO user_TableLastSaveString(string_ExcelImportPath, string_ExcelExportPath, string_ExcelColumnStart, string_ExcelColumnEnd, string_ExcelRowStart, string_ExcelRowEnd, string_ExcelSheetName) VALUES (?,?,?,?,?,?,?)', (self.root.ids.DataBind_FilePath.text, self.root.ids.DataBind_FileExportPath.text, self.root.ids.DataBind_ColumnCellStartPoint.text, self.root.ids.DataBind_ColumnCellEndPoint.text, self.root.ids.DataBind_RowCellStartPoint.text, self.root.ids.DataBind_RowCellEndPoint.text, self.root.ids.DataBind_ExcelFileSheet.text))

            SQLData.commit()
            SQLData.close()
            self.MDUserNotif_SnackbarHandler('Welcome, {0} {1}, | {2} |'.format(self.root.ids.FirstTimer_DataFirstName.text, self.root.ids.FirstTimer_DataLastName.text , self.root.ids.FirstTimer_DataJobPosition.text))
            self.MDNavDrawer_UserDefinition()
            self.MDScreenManagement_CheckVal()
    """
    (Callable Method, Function) => def UserIdentity_Edit(self), Edits User Information
        This one uses SQL Database and Assign Data from the Text Fields. Modifies SQL Database as well...
    """
    def UserIdentity_Edit(self, ExecutionType):
        DataIdentity = [Element for Element in range(4)] # For Storing User Identity
        user_OldFirstName = ''
        user_OldLastName = ''
        user_OldJobPosition =''
        Iterator = 1
        if ExecutionType == 'Receive':
            SQLData_Check = sqlite3.connect(self.SQLite_MainPath)
            SQLData_Cursor = SQLData_Check.cursor()
            SQLData_Cursor.execute('SELECT * FROM user_TableDatabase')
            SQLData_Fetch = SQLData_Cursor.fetchone()
            self.root.ids.UserEdit_FirstName.text = SQLData_Fetch[0]
            self.root.ids.UserEdit_LastName.text = SQLData_Fetch[1]
            self.root.ids.UserEdit_JobPosition.text = SQLData_Fetch[2]
            user_OldFirstName = SQLData_Fetch[0]
            user_OldLastName = SQLData_Fetch[1]
            user_OldJobPosition = SQLData_Fetch[2]
            SQLData_Check.close()
        elif ExecutionType == 'Submit':
            SQLData_Check = sqlite3.connect(self.SQLite_MainPath)
            SQLData_Cursor = SQLData_Check.cursor()
            SQLData_Cursor.execute('UPDATE user_TableDatabase SET user_FirstName = ?, user_LastName = ?, user_JobCurrent = ?', (self.root.ids.UserEdit_FirstName.text, self.root.ids.UserEdit_LastName.text, self.root.ids.UserEdit_JobPosition.text))
            SQLData_Check.commit()
            self.MDUserNotif_SnackbarHandler('User Data Updated!')
            self.root.ids.DataManagement_SectionAccess.disabled = False
            self.root.ids.Edit_SectionAccess.disabled = False
            SQLData_Check.close()
            self.MDNavDrawer_UserDefinition()

        #except:
        #    DataIdentity[0] = 'Unknown'
        #    DataIdentity[1] = 'User'
        #    DataIdentity[2] = 'Unknown'
        #self.root.ids.nav_drawer.header_name = 'Hello, {0} {1}!'. format(DataIdentity[0], DataIdentity[1])
        #self.root.ids.nav_drawer.subheader_name = 'Position: {0}'.format(DataIdentity[2])
        #pass
    """
    (Callable Method, Function) => def MDCard_Generate_DataStatusReport(self), Checkpoint that adds Specific Data to the Cards to Display Average and Totals of Other Such
        This function is the same as loading data from ExcelLoadInit_Data_Editor but it loads Data_only = True.
        It was seperated from that function because of the possible errors and trial recheck that could happen. So it's best to make button
        to generate that report.
    """
    def MDCard_Generate_DataStatusReport(self):
        self.DataReport_TotalUnits = 0
        self.DataReport_CalculatedProposed = 0
        self.DataReport_ActiveOnHand = 0
        self.DataReport_AverageOnHand = 0
        self.DataReport_AverageQuarterAllotment_1 = 0
        self.DataReport_AverageQuarterAllotment_2 = 0
        self.DataReport_AverageQuarterAllotment_3 = 0
        self.DataReport_AverageQuarterAllotment_4 = 0
        self.DataReport_AverageQuarterAllotment = 0
        self.DataReport_Calculated_TotalCost = 0
        self.DataReport_Calculated_UnitCost = 0
        """
            Reference on Calculation_TotalUnits Variable
            Issue: Iterating through elements that is composed as formulae don't update its values from the cache. Openpyxl doesn't evaluate formulas.
            So that concludes that the values from the cell contains None when loop through because it doesnt have cache value. So to do
            workarounds, we have to use win32.clint library to Open Excel and save it then close it without showing window. This was used so that
            Excel saves its value from the calculation of the formulae and when loop through, it has value that is None but integer or float.
            But furthermore, new discovery states that the cache is inaccurate. So inorder to do workarounds, We must compute for Total during
            iteration of the cells. Please sse calculation at Content_ListElement == 11. Check for others as you might see the data is binded when hit.
        """
        Calculation_Proposed = 0
        Calculation_Units = 0
        Content_ListElement = 0
        Content_CompletionList = 1
        TotalOnHand = 0
        try:
            ExcelFile = load_workbook(self.root.ids.DataBind_FilePath.text, data_only=True)
            ExcelWorksheet = ExcelFile['{0}'.format(self.root.ids.DataBind_ExcelFileSheet.text)]
            for EachDataSheet in ExcelWorksheet.iter_rows(min_row=int(self.root.ids.DataBind_RowCellStartPoint.text), max_row=int(self.root.ids.DataBind_RowCellEndPoint.text), min_col=column_index_from_string(self.root.ids.DataBind_ColumnCellStartPoint.text), max_col=column_index_from_string(self.root.ids.DataBind_ColumnCellEndPoint.text)):
                for cell in EachDataSheet:
                    Content_ListElement += 1
                    print('[Generate Report from Excel] Loop on Current Object', Content_ListElement, cell.value)
                    if Content_ListElement == 1 and Content_ListElement <= 15:
                        try:
                            self.DataReport_TotalUnits += 1 # Increment Only, because we are counting items here, not the value itself.s
                        except:
                            self.DataReport_TotalUnits += 0
                    elif Content_ListElement == 5 and Content_ListElement <= 15:
                        try:
                            self.DataReport_AverageOnHand += cell.value
                            self.DataReport_ActiveOnHand += 1
                            TotalOnHand += 1
                        except:
                            self.DataReport_ActiveOnHand += 0
                            TotalOnHand += 0
                    elif Content_ListElement == 6 and Content_ListElement <= 15:
                            continue
                    elif Content_ListElement == 7 and Content_ListElement <= 15:
                        try:
                            Calculation_Proposed = cell.value
                            self.DataReport_CalculatedProposed += cell.value
                        except:
                            Calculation_Proposed = 0
                            self.DataReport_CalculatedProposed += 0
                    elif Content_ListElement == 8 and Content_ListElement <= 15:
                            continue
                    elif Content_ListElement == 9 and Content_ListElement <= 15:
                            continue
                    elif Content_ListElement == 10 and Content_ListElement <= 15:
                        try:
                            Calculation_Units = cell.value
                            self.DataReport_Calculated_UnitCost += cell.value
                        except:
                            Calculation_Units = 0
                            self.DataReport_Calculated_UnitCost += 0.00
                    elif Content_ListElement == 11 and Content_ListElement <= 15:
                        try:
                            self.DataReport_Calculated_TotalCost += float(Calculation_Proposed * Calculation_Units)
                            Calculation_Proposed = 0
                            Calculation_Units = 0
                        except:
                            self.DataReport_Calculated_TotalCost += 0.00
                    elif Content_ListElement == 12 and Content_ListElement <= 15:
                        try:
                            self.DataReport_AverageQuarterAllotment_1 += cell.value
                        except:
                            self.DataReport_AverageQuarterAllotment_1 += 0
                    elif Content_ListElement == 13 and Content_ListElement <= 15:
                        try:
                            self.DataReport_AverageQuarterAllotment_2 += cell.value
                        except:
                            self.DataReport_AverageQuarterAllotment_2 += 0
                    elif Content_ListElement == 14 and Content_ListElement <= 15:
                        try:
                            self.DataReport_AverageQuarterAllotment_3 += cell.value
                        except:
                            self.DataReport_AverageQuarterAllotment_3 += 0
                    elif Content_ListElement == 15 and Content_ListElement <= 15:
                        try:
                            Content_ListElement = 0
                        except:
                            self.DataReport_AverageQuarterAllotment_4 += cell.value
                            Content_ListElement = 0
            self.root.ids.DataReport_TotalUnits.text = str(self.DataReport_TotalUnits)
            self.root.ids.DataReport_TotalProposed.text = str(self.DataReport_CalculatedProposed / TotalOnHand)
            self.root.ids.DataReport_TotalOnHand.text = str(self.DataReport_ActiveOnHand)
            self.root.ids.DataReport_CalculatedOnHand.text = str(float(self.DataReport_AverageOnHand / TotalOnHand))
            self.root.ids.DataReport_TotalCost.text = str(self.DataReport_Calculated_TotalCost)
            self.root.ids.DataReport_UnitCost.text = str(self.DataReport_Calculated_UnitCost)
            self.root.ids.DataReport_QuarterAllotment.text = str(float(self.DataReport_AverageQuarterAllotment_1 + self.DataReport_AverageQuarterAllotment_2 + self.DataReport_AverageQuarterAllotment_3 + self.DataReport_AverageQuarterAllotment_4) / 4)
        except Exception as ErrorMessage:
            self.MDUserNotif_SnackbarHandler('Error, cannot generate report due to failed operation on calculation.')
            self.MDUserNotif_SnackbarHandler(str(ErrorMessage))
    """
    (Callable Method, Function) => def MDUserNotif_SnackbarHandler(self), Checkpoint, Shows Status or Shows Notification about the specific Function.
    params_message has no limit as far as we know, not sure how MAX_PATH will affect this one if characters are more than 255.
    """
    def MDUserNotif_SnackbarHandler(self, params_message):
        Snackbar(text=params_message).show()
        
    """
    FUNCTION RUNTIME MODIFIERS SECTIONS
        It is a set of functions that modifies properties of the program such as design.
        Excel is not part of this section due to the fact that is involves data.

    (Callable Method, Function) => def MDToolbar_DynamicContent_TitleChange(self), The Window Identifier
        This function modifies the Toolbar Title Text by Sending specific strings to process by calling this function.
    """
    def MDToolbar_DynamicContent_TitleChange(self, params_title):
        self.root.ids.ToolbarMain.title = params_title

    """
    (Callable Method, Function) => def GUI_ColorChange(self), Change GUI Color on Runtime due to Bad Color Mix
        This function is usable on click on navigation drawer. It reuse part of Method Non Collable, Function build(self)
    """
    def GUI_ColorChange(self):
        rand_primary_palette = ['Pink', 'Blue', 'Indigo', 'BlueGrey', 'Brown','LightBlue','Purple', 'Yellow', 'DeepOrange','Green', 'Red', 'Teal', 'Orange', 'Cyan', 'Amber','DeepPurple', 'Lime']
        self.theme_cls.primary_palette = (RandomizePick(rand_primary_palette))
    """
    (Callable Method, Function) => def MDButton_DarkMode(self), Light and Dark Mode, Sets Whole Widgets to Specific Data Given
        This function is usable on click of the button on top right. It changes texture color based on the data given. It is logical
        so no futher descriptions should be added.
    """
    def MDButton_DarkMode(self):
        if self.root.ids.ToolbarMain.darkBooleanParameter == True and self.theme_cls.theme_style == 'Dark':
            self.root.ids.ToolbarMain.darkBooleanParameter = False
            self.root.ids.ToolbarMain.right_action_items =  [['invert-colors', lambda x: MD_InventoryEXDI_SMLV_GUI.GUI_ColorChange()], ['brightness-2', lambda x: MD_InventoryEXDI_SMLV_GUI.MDButton_DarkMode(self)]]
            self.theme_cls.theme_style = 'Light'
            # Make it time based, if possible
        elif self.root.ids.ToolbarMain.darkBooleanParameter == False and self.theme_cls.theme_style == 'Light':
            self.root.ids.ToolbarMain.darkBooleanParameter = True
            self.root.ids.ToolbarMain.right_action_items =  [['invert-colors', lambda x: MD_InventoryEXDI_SMLV_GUI.GUI_ColorChange()], ['brightness-7', lambda x: MD_InventoryEXDI_SMLV_GUI.MDButton_DarkMode(self)]]
            self.theme_cls.theme_style = 'Dark'

    """ 
    EXCEL DATA LOADING SECTION
        A set of functions that uses KV Widgets to display Data using openpyxl or even manipulate data
    
    (Callable Method, Function) => def ExcelLoadInit_Data_Editor(self), Loads Data to MDList on the Left Side of the Data Management 
    x Three Arguments are given in order to run this function: 
    Unload -> Removes all data from the MDList and TextField as well, disabled some buttons to indicate that there is no data to do something based on the button description
    load -> Loads all data based on parameters set on Import Data / Export Pull Out, see logic ExcelLoadInit_Data_Editor for more info.
    Reload -> Unloads and Loads the data. Descriptions of each is the same as from the above.
    LOGIC
    """
    def ExcelLoadInit_Data_Editor(self, ExcelLoad_Method):
        """
            Initialize Variables, some of these variables are used for concatenation of the data or even conversions if there is any.
            Those variables with null strings are something that we called a placeholder for data processing.
        """
        Content_ListElement = 0
        Content_CompletionList = 1
        ExcelContainer_Val_ItemNumber = ''
        ExcelContainer_Text_ParticularName = ''
        ExcelContainer_Val_OnHand = ''
        ExcelContainer_Val_Proposed = ''
        ExcelContainer_CostUnit_Type = ''
        ExcelContainer_CostUnit_Val = ''
        ExcelContainer_CostUnit_Total = ''
        ExcelContainer_Quarter_Allot_1 = ''
        ExcelContainer_Quarter_Allot_2 = ''
        ExcelContainer_Quarter_Allot_3 = ''
        ExcelContainer_Quarter_Allot_4 = ''
        Concatenation_PrimaryText = ''
        Concatenation_SecondaryText = ''
        Start_CellPoint = ''
        End_CellPoint = ''
        try:
            """
                Load Data from Excel: Here, we try to load the data based on the parameters given, all parameters here must be precised and it is Case Sensitive.
            """
            ExcelFile = load_workbook(self.root.ids.DataBind_FilePath.text)
            ExcelWorksheet = ExcelFile['{0}'.format(self.root.ids.DataBind_ExcelFileSheet.text)]
            """
                Unloading Data from GUI: We iterate through 100 since the Maximum Columns of the MDList that can naturally handle is 100.
                Per each element or per each list item, data will be set to NULL or seomthing that shows "No Data"
            """
            if ExcelLoad_Method == 'Unload' or ExcelLoad_Method == 'Reload':
                for Element_InList in range(1, 101):
                    print('[Excel Data Unload / Reload] Removing Data from ListElement # ', Element_InList)
                    self.root.ids['ListElement_%d' % Element_InList].cellStart = '0'
                    self.root.ids['ListElement_%d' % Element_InList].cellEnd = '0'
                    self.root.ids['ListElement_%d' % Element_InList].text = '#??? - <No Data>'
                    self.root.ids['ListElement_%d' % Element_InList].secondary_text = 'There is no data to load...'
                    self.root.ids['ListElement_%d' % Element_InList].disabled = True
                    self.root.ids['ListElement_%d' % Element_InList].itemNumber = '0'
                    self.root.ids['ListElement_%d' % Element_InList].particularName = '0'
                    self.root.ids['ListElement_%d' % Element_InList].onHandVal = '0'
                    self.root.ids['ListElement_%d' % Element_InList].proposedVal = '0'
                    self.root.ids['ListElement_%d' % Element_InList].unitType = '0'
                    self.root.ids['ListElement_%d' % Element_InList].unitVal = '0'
                    self.root.ids['ListElement_%d' % Element_InList].quarterOne = '0'
                    self.root.ids['ListElement_%d' % Element_InList].quarterTwo = '0'
                    self.root.ids['ListElement_%d' % Element_InList].quarterThree = '0'
                    self.root.ids['ListElement_%d' % Element_InList].quarterFour = '0'
                    self.root.ids['ListElement_%d' % Element_InList].totalVal = '0'
                self.root.ids.DataReport_TotalUnits.text  = ''
                self.root.ids.DataReport_TotalProposed.text = ''
                self.root.ids.DataReport_TotalOnHand.text = ''
                self.root.ids.DataReport_CalculatedOnHand.text = ''
                self.root.ids.DataReport_TotalCost.text = ''
                self.root.ids.DataReport_UnitCost.text = ''
                self.root.ids.DataReport_QuarterAllotment.text = ''
                self.DataReport_TotalUnits = ''
                self.DataReport_Calculated_TotalCost = ''
                self.DataReport_CalculatedProposed = ''
                self.DataReport_Calculated_UnitCost = ''
                self.DataReport_ActiveOnHand = ''
                self.DataReport_AverageOnHand = ''
                self.DataReport_AverageQuarterAllotment_1 = ''
                self.DataReport_AverageQuarterAllotment_2 = ''
                self.DataReport_AverageQuarterAllotment_3 = ''
                self.DataReport_AverageQuarterAllotment_4 = ''
                self.root.ids.Resource_ItemNumVal.text = ''
                self.root.ids.Resource_ParticularProperty.text = ''
                self.root.ids.Resource_OnHandVal.text = ''
                self.root.ids.Resource_ProposedVal.text = ''
                self.root.ids.Resource_UnitTypeProperty.text = ''
                self.root.ids.Resource_UnitVal.text = ''
                self.root.ids.Quartile_TextFieldVal_One.text = ''
                self.root.ids.Quartile_TextFieldVal_Two.text = ''
                self.root.ids.Quartile_TextFieldVal_Three.text = ''
                self.root.ids.Quartile_TextFieldVal_Four.text = ''
                self.root.ids.Total_ComputedVal.text = ''
                self.root.ids.FileOpt_Progress_SaveImport.disabled = True
                self.root.ids.FileOpt_Progress_SaveExport.disabled = True
                self.root.ids.FileOpt_ReInit_ReloadDataFile.disabled = True
                self.root.ids.FileOpt_EndInst_UnloadDataFile.disabled = True
                self.root.ids.FileOpt_General_GenerateReport.disabled = True
                self.root.ids.FileOpt_FirstInst_LoadDataFile.disabled = False
                self.root.ids.MDDataEditor_ApplyChange.disabled = True
                self.root.ids.MDDataEditor_ClearInput.disabled = True
                self.Active_DataElement = ''
                self.MDToolbar_DynamicContent_TitleChange(self.title)
                self.MDUserNotif_SnackbarHandler("Binded Data has been removed from the Editor.")
                """
                    Loading Data from Excel to GUI: This is where, we are going to use the load_workbook variable holder.
                    .iter_rows(needs min_row, max_row, min_col and max_col inroder to run, it also says where to start reading the data and where to stop)
                    Function from openpyxl ~ column_index_from_string converts the parameter set into a number, thats because iter_rows only accepts number
                    so in order to be user friendly and avoiding disadvantage, we do workarounds and that is the function that helps us throughout this problem
                    
                    NOTE: I can create those widgets dynamically but not with IDS by Janrey Licas
                        I would rather go to Iterate Through IDS of the Statically Created ListItem Method than
                        trying to use weakref and other such methods when in fact I will just get myself go to the next problem such as 
                        accessing ids in runtime without updating self.ids... It is not updatable... Since kivy store those ids in self.ids without update
                        even adding a wdiget. So since, the project deadlien is almost ahead of time. We would rather go to this static method, it's really hard to accept
                        as a Lead Developer of this application but since time goes by, we have to implement something even though this is one of the most important component in the 
                        entire application. In the recent template, there are spcaes that needs to be avoided...
                """
            if ExcelLoad_Method == 'Load' or ExcelLoad_Method == 'Reload':
                for EachDataSheet in ExcelWorksheet.iter_rows(min_row=int(self.root.ids.DataBind_RowCellStartPoint.text), max_row=int(self.root.ids.DataBind_RowCellEndPoint.text), min_col=column_index_from_string(self.root.ids.DataBind_ColumnCellStartPoint.text), max_col=column_index_from_string(self.root.ids.DataBind_ColumnCellEndPoint.text)):
                    for cell in EachDataSheet:
                        """
                        For each cell, it has it's own place (according to clients given template) we have Content_ListElement to help us place that value from respective data placment.
                        For each Content_ListElement, there is some checks that is happening, that is meant to be happen so that if some data seems to be misleading, we have to replace it
                        with something to indicate that this data is empty or ready to be filled up.
                        """
                        Content_ListElement += 1
                        print('[Binding Excel Data to] Loop On Current Object', Content_ListElement)
                        if Content_ListElement == 1 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_Val_ItemNumber = 0
                                Start_CellPoint = '{0}{1}'.format(cell.column, cell.row)
                                #self.root.ids['ListElement_%d' % Content_CompletionList].cellStart = Start_CellPoint
                            else:
                                ExcelContainer_Val_ItemNumber = cell.value
                                Start_CellPoint = '{0}{1}'.format(cell.column, cell.row)
                                #self.root.ids['ListElement_%d' % Content_CompletionList].cellStart = Start_CellPoint
                        elif Content_ListElement == 2 and Content_ListElement <= 15:
                                continue
                        elif Content_ListElement == 3 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_Text_ParticularName = 'No Particular Description'
                            else:
                                ExcelContainer_Text_ParticularName = cell.value
                        elif Content_ListElement == 4 and Content_ListElement <= 15:
                                continue
                        elif Content_ListElement == 5 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_Val_OnHand = 0
                            else:
                                ExcelContainer_Val_OnHand = cell.value
                        elif Content_ListElement == 6 and Content_ListElement <= 15:
                                continue
                        elif Content_ListElement == 7 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_Val_Proposed = 0
                            else:
                                ExcelContainer_Val_Proposed = cell.value
                        elif Content_ListElement == 8 and Content_ListElement <= 15:
                                continue
                        elif Content_ListElement == 9 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_CostUnit_Type = 'Unknown'
                            else:
                                ExcelContainer_CostUnit_Type = cell.value
                        elif Content_ListElement == 10 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_CostUnit_Val = 0.00
                            else:
                                ExcelContainer_CostUnit_Val = cell.value
                        elif Content_ListElement == 11 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_CostUnit_Total = 0.00
                            else:
                                ExcelContainer_CostUnit_Total = cell.value
                        elif Content_ListElement == 12 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_Quarter_Allot_1 = 0
                            else:
                                ExcelContainer_Quarter_Allot_1 = cell.value
                        elif Content_ListElement == 13 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_Quarter_Allot_2 = 0
                            else:
                                ExcelContainer_Quarter_Allot_2 = cell.value
                        elif Content_ListElement == 14 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_Quarter_Allot_3 = 0
                            else:
                                ExcelContainer_Quarter_Allot_3 = cell.value
                        elif Content_ListElement == 15 and Content_ListElement <= 15:
                            if cell.value is None or cell.value == '' or cell.value == '-':
                                ExcelContainer_Quarter_Allot_4 = 0
                                Content_ListElement = 0
                                End_CellPoint = '{0}{1}'.format(cell.column, cell.row)
                                #self.root.ids['ListElement_%d' % Content_CompletionList].cellEnd = End_CellPoint
                                break
                            else:
                                ExcelContainer_Quarter_Allot_4 = cell.value
                                Content_ListElement = 0
                                End_CellPoint = '{0}{1}'.format(cell.column, cell.row)
                                #self.root.ids['ListElement_%d' % Content_CompletionList].cellEnd = End_CellPoint
                                break
                    """
                        We concatenate those variables into one text and process it under the corresponding textfields.
                        We dont do direct approach due to some testing, some parts are missing or wasn't addressed to the point it doesnt display those values under the MDListItem.
                    """
                    self.root.ids['ListElement_%d' % Content_CompletionList].text = '{0} | {1}'.format(ExcelContainer_Val_ItemNumber,ExcelContainer_Text_ParticularName)
                    self.root.ids['ListElement_%d' % Content_CompletionList].secondary_text = 'OnHand: {0}, Proposed:{1}, Unit Type:{2}, Unit Value:{3}, TotalVal:{4}, Q1:{5}, Q2:{6}, Q3:{7}, Q4:{8}'.format(ExcelContainer_Val_OnHand, ExcelContainer_Val_Proposed, ExcelContainer_CostUnit_Type, ExcelContainer_CostUnit_Val, ExcelContainer_CostUnit_Total, ExcelContainer_Quarter_Allot_1, ExcelContainer_Quarter_Allot_2, ExcelContainer_Quarter_Allot_3, ExcelContainer_Quarter_Allot_4)
                    self.root.ids['ListElement_%d' % Content_CompletionList].disabled = False
                    """
                    Save those values as well in specific property instead of just takes. Thats because we cannot get that text from the string itself.
                    Its best to design its own container in order to get the results with efficient.
                    """
                    self.root.ids['ListElement_%d' % Content_CompletionList].itemNumber = str(ExcelContainer_Val_ItemNumber)
                    self.root.ids['ListElement_%d' % Content_CompletionList].particularName = str(ExcelContainer_Text_ParticularName)
                    self.root.ids['ListElement_%d' % Content_CompletionList].onHandVal = str(ExcelContainer_Val_OnHand)
                    self.root.ids['ListElement_%d' % Content_CompletionList].proposedVal = str(ExcelContainer_Val_Proposed)
                    self.root.ids['ListElement_%d' % Content_CompletionList].unitType = str(ExcelContainer_CostUnit_Type)
                    self.root.ids['ListElement_%d' % Content_CompletionList].unitVal = str(ExcelContainer_CostUnit_Val)
                    self.root.ids['ListElement_%d' % Content_CompletionList].quarterOne = str(ExcelContainer_Quarter_Allot_1)
                    self.root.ids['ListElement_%d' % Content_CompletionList].quarterTwo = str(ExcelContainer_Quarter_Allot_2)
                    self.root.ids['ListElement_%d' % Content_CompletionList].quarterThree = str(ExcelContainer_Quarter_Allot_3)
                    self.root.ids['ListElement_%d' % Content_CompletionList].quarterFour = str(ExcelContainer_Quarter_Allot_4)
                    self.root.ids['ListElement_%d' % Content_CompletionList].totalVal = str(ExcelContainer_CostUnit_Total)
                    Content_CompletionList += 1 # Increment by 1 because we are done to this list so, move on to the next list.
                Current_ActiveContentTitle = ExcelWorksheet['F%d' % (int(self.root.ids.DataBind_RowCellStartPoint.text) - 4)].value
                try:
                    Formatted_ContentTitle = Current_ActiveContentTitle.lstrip()
                    self.MDToolbar_DynamicContent_TitleChange('EXDI_SMLV | {0}'.format(Formatted_ContentTitle))
                except:
                    self.MDToolbar_DynamicContent_TitleChange('EXDI_SMLV | Unknown')
                self.MDUserNotif_SnackbarHandler("File Data Load Success!")
                if ExcelLoad_Method == 'Load':
                    self.root.ids.FileOpt_Progress_SaveImport.disabled = False
                    self.root.ids.FileOpt_Progress_SaveExport.disabled = False
                    self.root.ids.FileOpt_FirstInst_LoadDataFile.disabled = True
                    self.root.ids.FileOpt_ReInit_ReloadDataFile.disabled = False
                    self.root.ids.FileOpt_EndInst_UnloadDataFile.disabled = False
                    self.root.ids.FileOpt_General_GenerateReport.disabled = False
                    self.root.ids.MDDataEditor_ApplyChange.disabled = False
                    self.root.ids.MDDataEditor_ClearInput.disabled = False
                elif ExcelLoad_Method == 'Reload':
                    self.root.ids.FileOpt_Progress_SaveImport.disabled = False
                    self.root.ids.FileOpt_Progress_SaveExport.disabled = False
                    self.root.ids.FileOpt_FirstInst_LoadDataFile.disabled = False
                    self.root.ids.FileOpt_ReInit_ReloadDataFile.disabled = False
                    self.root.ids.FileOpt_EndInst_UnloadDataFile.disabled = False
                    self.root.ids.FileOpt_General_GenerateReport.disabled = False
                    self.root.ids.MDDataEditor_ApplyChange.disabled = False
                    self.root.ids.MDDataEditor_ClearInput.disabled = False
        except Exception as ErrorMessage:
            self.MDUserNotif_SnackbarHandler("Failed to load data... Check the files or the arguments you set on the Data Import Data / Export Pull Out Section!")
            self.MDUserNotif_SnackbarHandler(str(ErrorMessage))
    """
        (Callable Method, Function) => def ExcelLoadDataTextField_OnClick(self), Property Gatherer for each ListItem
            Loads data to Respective TextFeilds in order to start editing the data.
    """
    def ExcelLoadDataTextField_OnClick(self, ListElement_Reference):
        self.root.ids.Resource_ItemNumVal.text = self.root.ids['%s' % ListElement_Reference].itemNumber
        self.root.ids.Resource_ParticularProperty.text = self.root.ids['%s' % ListElement_Reference].particularName
        self.root.ids.Resource_OnHandVal.text = self.root.ids['%s' % ListElement_Reference].onHandVal
        self.root.ids.Resource_ProposedVal.text = self.root.ids['%s' % ListElement_Reference].proposedVal
        self.root.ids.Resource_UnitTypeProperty.text = self.root.ids['%s' % ListElement_Reference].unitType
        self.root.ids.Resource_UnitVal.text = self.root.ids['%s' % ListElement_Reference].unitVal
        self.root.ids.Quartile_TextFieldVal_One.text = self.root.ids['%s' % ListElement_Reference].quarterOne
        self.root.ids.Quartile_TextFieldVal_Two.text = self.root.ids['%s' % ListElement_Reference].quarterTwo
        self.root.ids.Quartile_TextFieldVal_Three.text = self.root.ids['%s' % ListElement_Reference].quarterThree
        self.root.ids.Quartile_TextFieldVal_Four.text = self.root.ids['%s' % ListElement_Reference].quarterFour
        self.root.ids.Total_ComputedVal.text = self.root.ids['%s' % ListElement_Reference].totalVal
        self.Active_DataElement = ListElement_Reference
    """
        (Callable Method, Function) => def MDWorksheetWorker_DataEditor_ValueRaiseLowCaller(self), Increment / Decrement Value by checking its string if thats a integer or float.
            Converts string to integer, and does increment or decrement based on Passed Argument @ TypeAction_ValueChange
            What this method cannot do? Float Incrementation or Decrementation. Cannot execute when the textfields contains dot representing there is a float value
    """
    def MDWorksheetWorker_DataEditor_ValueRaiseLowCaller(self, Button_IDNumber, TypeAction_ValueChange):
        try:
            if Button_IDNumber == 1:
                if TypeAction_ValueChange == 'Increment':
                    try:
                        self.root.ids.Resource_ItemNumVal.text = str(int(self.root.ids.Resource_ItemNumVal.text) + 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot increment this value. (Was this value a float or string?)')
                elif TypeAction_ValueChange == 'Decrement':
                    try:
                        self.root.ids.Resource_ItemNumVal.text = str(int(self.root.ids.Resource_ItemNumVal.text) - 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot decrement this value. (Was this value a float or string?)')
                else:
                    raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
            elif Button_IDNumber == 2:
                if TypeAction_ValueChange == 'Increment':
                    try:
                        self.root.ids.Resource_OnHandVal.text = str(int(self.root.ids.Resource_OnHandVal.text) + 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot increment this value. (Was this value a float or string?)')
                elif TypeAction_ValueChange == 'Decrement':
                    try:
                        self.root.ids.Resource_OnHandVal.text = str(int(self.root.ids.Resource_OnHandVal.text) - 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot decrement this value. (Was this value a float or string?)')
                else:
                    raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
            elif Button_IDNumber == 3:
                if TypeAction_ValueChange == 'Increment':
                    try:
                        self.root.ids.Resource_ProposedVal.text = str(int(self.root.ids.Resource_ProposedVal.text) + 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot increment this value. (Was this value a float or string?)')
                elif TypeAction_ValueChange == 'Decrement':
                    try:
                        self.root.ids.Resource_ProposedVal.text = str(int(self.root.ids.Resource_ProposedVal.text) - 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot decrement this value. (Was this value a float or string?)')
                else:
                    raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
            elif Button_IDNumber == 4:
                if TypeAction_ValueChange == 'Increment':
                    try:
                        self.root.ids.Resource_UnitVal.text = str(int(self.root.ids.Resource_UnitVal.text) + 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot increment this value. (Was this value a float or string?)')
                elif TypeAction_ValueChange == 'Decrement':
                    try:
                        self.root.ids.Resource_UnitVal.text = str(int(self.root.ids.Resource_UnitVal.text) - 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot decrement this value. (Was this value a float or string?)')
                else:
                    raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
            elif Button_IDNumber == 5:
                if TypeAction_ValueChange == 'Increment':
                    try:
                        self.root.ids.Quartile_TextFieldVal_One.text  = str(int(self.root.ids.Quartile_TextFieldVal_One.text) + 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot increment this value. (Was this value a float or string?)')
                elif TypeAction_ValueChange == 'Decrement':
                    try:
                        self.root.ids.Quartile_TextFieldVal_One.text  = str(int(self.root.ids.Quartile_TextFieldVal_One.text) - 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot decrement this value. (Was this value a float or string?)')
                else:
                    raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
            elif Button_IDNumber == 6:
                if TypeAction_ValueChange == 'Increment':
                    try:
                        self.root.ids.Quartile_TextFieldVal_Two.text = str(int(self.root.ids.Quartile_TextFieldVal_Two.text) + 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot increment this value. (Was this value a float or string?)')
                elif TypeAction_ValueChange == 'Decrement':
                    try:
                        self.root.ids.Quartile_TextFieldVal_Two.text  = str(int(self.root.ids.Quartile_TextFieldVal_Two.text) - 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot decrement this value. (Was this value a float or string?)')
                else:
                    raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
            elif Button_IDNumber == 7:
                if TypeAction_ValueChange == 'Increment':
                    try:
                        self.root.ids.Quartile_TextFieldVal_Three.text = str(int(self.root.ids.Quartile_TextFieldVal_Three.text) + 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot increment this value. (Was this value a float or string?)')
                elif TypeAction_ValueChange == 'Decrement':
                    try:
                        self.root.ids.Quartile_TextFieldVal_Three.text = str(int(self.root.ids.Quartile_TextFieldVal_Three.text) - 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot decrement this value. (Was this value a float or string?)')
                else:
                    raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
            elif Button_IDNumber == 8:
                if TypeAction_ValueChange == 'Increment':
                    try:
                        self.root.ids.Quartile_TextFieldVal_Four.text = str(int(self.root.ids.Quartile_TextFieldVal_Four.text) + 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot increment this value. (Was this value a float or string?)')
                elif TypeAction_ValueChange == 'Decrement':
                    try:
                        self.root.ids.Quartile_TextFieldVal_Four.text = str(int(self.root.ids.Quartile_TextFieldVal_Four.text) - 1)
                    except:
                        self.MDUserNotif_SnackbarHandler('Sorry, I cannot decrement this value. (Was this value a float or string?)')
                else:
                    raise ValueError('MDWorksheetWorker_DataEditor_ValueRaiseLowCaller Error: Data Passed is unable to met conditions with the code.')
        except:
            self.MDUserNotif_SnackbarHandler('Sorry, the textinput value is not valid for incrementing or decrementing data.')
    
    """
        (Callable Method, Function) => def ExcelFile_SaveOnPath(self), Saves ExcelFile, Location Dependent on External
        Two Types of Arguments can be passed: CurrentPath and ExportPath
        This function uses the same method of pulling data from excel to GUI. But this time, it will GUI to Excel
        Please refer to (Callable Method, Function) ExcelLoadInit_Data_Editor() on how it works.
    """
    def ExcelFile_SaveOnPath(self, SavePath_Type):
        ExcelFile = load_workbook(self.root.ids.DataBind_FilePath.text, data_only=False)
        ExcelWorksheet_Properties = ExcelFile.properties
        ExcelWorksheet_Active = ExcelFile.active
        ExcelWorksheet_Init = Workbook()
        Content_ListElement = 0
        Content_CompletionList = 1
        try:
            ExcelWorksheet_Properties.creator = '{0} {1} | {2}'.format(self.DataIdentity[0], self.DataIdentity[1], self.DataIdentity[2])
            ExcelWorksheet_Properties.lastModifiedBy = '{0} {1}| {2}'.format(self.DataIdentity[0], self.DataIdentity[1], self.DataIdentity[2])
            ExcelWorksheet_Properties.last_modified_by = '{0} {1} | {2}'.format(self.DataIdentity[0], self.DataIdentity[1], self.DataIdentity[2])
            for EachDataSheet in ExcelWorksheet_Active.iter_rows(min_row=int(self.root.ids.DataBind_RowCellStartPoint.text), max_row=int(self.root.ids.DataBind_RowCellEndPoint.text), min_col=column_index_from_string(self.root.ids.DataBind_ColumnCellStartPoint.text), max_col=column_index_from_string(self.root.ids.DataBind_ColumnCellEndPoint.text)):
                for cell in EachDataSheet:
                    Content_ListElement += 1
                    print('[Overwrite Data to Excel] Element Content', Content_ListElement)
                    if Content_ListElement == 1 and Content_ListElement <= 15:
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = int(self.root.ids['ListElement_%d' % Content_CompletionList].itemNumber)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].itemNumber)
                    elif Content_ListElement == 2 and Content_ListElement <= 15:
                        ExcelWorksheet_Active['%s' % cell.coordinate] = ''
                    elif Content_ListElement == 3 and Content_ListElement <= 15:
                        ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].particularName)
                    elif Content_ListElement == 4 and Content_ListElement <= 15:
                        ExcelWorksheet_Active['%s' % cell.coordinate] = ''
                    elif Content_ListElement == 5 and Content_ListElement <= 15:
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = int(self.root.ids['ListElement_%d' % Content_CompletionList].onHandVal)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].onHandVal)
                    elif Content_ListElement == 6 and Content_ListElement <= 15:
                        ExcelWorksheet_Active['%s' % cell.coordinate] = ''
                    elif Content_ListElement == 7 and Content_ListElement <= 15:
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = int(self.root.ids['ListElement_%d' % Content_CompletionList].proposedVal)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].proposedVal)
                    elif Content_ListElement == 8 and Content_ListElement <= 15:
                        ExcelWorksheet_Active['%s' % cell.coordinate] = ''
                    elif Content_ListElement == 9 and Content_ListElement <= 15:
                        ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].unitType)
                    elif Content_ListElement == 10 and Content_ListElement <= 15:
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = float(self.root.ids['ListElement_%d' % Content_CompletionList].unitVal)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].unitVal)
                    elif Content_ListElement == 11 and Content_ListElement <= 15:
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = float(self.root.ids['ListElement_%d' % Content_CompletionList].totalVal)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].totalVal)
                    elif Content_ListElement == 12 and Content_ListElement <= 15:
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = int(self.root.ids['ListElement_%d' % Content_CompletionList].quarterOne)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].quarterOne)
                    elif Content_ListElement == 13 and Content_ListElement <= 15:
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = int(self.root.ids['ListElement_%d' % Content_CompletionList].quarterTwo)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].quarterTwo)
                    elif Content_ListElement == 14 and Content_ListElement <= 15:
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = int(self.root.ids['ListElement_%d' % Content_CompletionList].quarterThree)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].quarterThree)
                    elif Content_ListElement == 15 and Content_ListElement <= 15:
                        Content_ListElement = 0
                        try:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = int(self.root.ids['ListElement_%d' % Content_CompletionList].quarterFour)
                        except:
                            ExcelWorksheet_Active['%s' % cell.coordinate] = str(self.root.ids['ListElement_%d' % Content_CompletionList].quarterFour)
                Content_CompletionList += 1
            """
                There is no difference, except its location based on parameters passed from the data import and export pull out
            """ 
            if SavePath_Type == 'CurrentPath':
                ExcelFile.save(self.root.ids.DataBind_FilePath.text)
                self.MDUserNotif_SnackbarHandler('Saved in Current Path! -> {0}'.format(self.root.ids.DataBind_FilePath.text))
            elif SavePath_Type == 'ExportPath':
                try:
                    if len(self.root.ids.DataBind_FileExportPath.text) >= 4:
                        ExcelFile.save(self.root.ids.DataBind_FileExportPath.text)
                        self.MDUserNotif_SnackbarHandler('Saved in Export Path! -> {0}'.format(self.root.ids.DataBind_FilePath.text))
                except Exception as ErrorMessage:
                    self.MDUserNotif_SnackbarHandler('Error, Export Path given is either Permission Denied or does not exist. Please select other path or check this next message.')
                    self.MDUserNotif_SnackbarHandler(str(ErrorMessage))
            else:
                raise ValueError('Argument Passed for SaveTypePath is not valid.')
        except Exception as ErrorMessage:
            self.MDUserNotif_SnackbarHandler('Error, one of the values may not be respective to float or integer. Or close the excel file that you are currently editing!')
            self.MDUserNotif_SnackbarHandler(str(ErrorMessage))
    """
    (Callable Method, Function) def MDWorksheetWorker_DataEditor_Apply, Applies Modified Data from TextField to MDList, Custom Properties receives from the TextInput Widget.
    There is a part that refreshes the list or the strings that user can see, it modifies text and secondary_text.
    Also this approach is direct, according to ExcelLoadInit_Data_Editor function, we didnt do direct approach due to some reasons. (Please go to that function if you want to learn more about that approach.)
    So in conclusion, it is possible. But hence, no changes from that function will be made.
    """
    def MDWorksheetWorker_DataEditor_Apply(self, ListElement_Change):
        try:
            self.root.ids['%s'% ListElement_Change].itemNumber = self.root.ids.Resource_ItemNumVal.text
            self.root.ids['%s'% ListElement_Change].particularName = self.root.ids.Resource_ParticularProperty.text
            self.root.ids['%s'% ListElement_Change].onHandVal = self.root.ids.Resource_OnHandVal.text
            self.root.ids['%s'% ListElement_Change].proposedVal = self.root.ids.Resource_ProposedVal.text
            self.root.ids['%s'% ListElement_Change].unitType = self.root.ids.Resource_UnitTypeProperty.text
            self.root.ids['%s'% ListElement_Change].unitVal = self.root.ids.Resource_UnitVal.text
            self.root.ids['%s'% ListElement_Change].quarterOne = self.root.ids.Quartile_TextFieldVal_One.text
            self.root.ids['%s'% ListElement_Change].quarterTwo = self.root.ids.Quartile_TextFieldVal_Two.text
            self.root.ids['%s'% ListElement_Change].quarterThree = self.root.ids.Quartile_TextFieldVal_Three.text
            self.root.ids['%s'% ListElement_Change].quarterFour = self.root.ids.Quartile_TextFieldVal_Four.text
            self.root.ids['%s'% ListElement_Change].totalVal = self.root.ids.Total_ComputedVal.text
            self.root.ids['%s' % ListElement_Change].text = '{0} | {1}'.format(self.root.ids.Resource_ItemNumVal.text,self.root.ids.Resource_ParticularProperty.text)
            self.root.ids['%s' % ListElement_Change].secondary_text = 'OnHand: {0}, Proposed:{1}, Unit Type:{2}, Unit Value:{3}, TotalVal:{4}, Q1:{5}, Q2:{6}, Q3:{7}, Q4:{8}'.format(self.root.ids.Resource_OnHandVal.text, self.root.ids.Resource_ProposedVal.text, self.root.ids.Resource_UnitTypeProperty.text, self.root.ids.Resource_UnitVal.text, self.root.ids.Total_ComputedVal.text, self.root.ids.Quartile_TextFieldVal_One.text, self.root.ids.Quartile_TextFieldVal_Two.text, self.root.ids.Quartile_TextFieldVal_Three.text, self.root.ids.Quartile_TextFieldVal_Four.text)
            if len(self.root.ids.DataReport_ResourceLastModified.text) >= 1:
                self.root.ids.DataReport_ResourceLastModified.text = self.root.ids['%s'% ListElement_Change].particularName
            else:
                self.root.ids.DataReport_ResourceLastModified.text = 'Unknown Particular'
            self.MDUserNotif_SnackbarHandler('Modified Data has been applied to selected list item!')
        except:
            self.MDUserNotif_SnackbarHandler('Sorry, you cannot do that!')

    """
    (Callable Method, Function) def IMD_EXPO_ButtonCallBackManager, Button Handler @ Data Editor and File Options Section
        Created this for a reason, we dont want to do more functions for the buttons as the UML will be overloaded. In order to optimize without thinking for a better approach
        Hence, we know this might be the best non-professional approach, we did one function, one arguments. This was intended to avoid documenting more functions.
        Most of this calls, are for clearing specific textfields that button has in relationship with only. Buttons with specific actions is not a good decision.
        That's because one functoin that has loads of content is not one of the best practical apporach as per Project Leader learned from other people.
    """
    def IMD_EXPO_ButtonCallBackManager(self, params_WidgetID):
        if params_WidgetID == 'ClearInput_Path':
            self.root.ids.DataBind_FilePath.text = ''
            self.MDUserNotif_SnackbarHandler('Excel File Path Input has been cleared!')
        elif params_WidgetID == 'CallFunc_ClearEditTextFields':
            self.root.ids.UserEdit_FirstName.text = ''
            self.root.ids.UserEdit_LastName.text = ''
            self.root.ids.UserEdit_JobPosition.text = ''
            self.root.ids.UserEdit_Password.text = ''
        elif params_WidgetID == 'CallFunc_ClearFirstTimeTextFields':
            self.root.ids.FirstTimer_DataFirstName.text = ''
            self.root.ids.FirstTimer_DataLastName.text = ''
            self.root.ids.FirstTimer_DataJobPosition.text = ''
            self.root.ids.FirstTimer_DataPassword.text = ''
        elif params_WidgetID == 'CallFunc_ClearColumnStartPoint':
            self.root.ids.DataBind_ColumnCellStartPoint.text = ''
        elif params_WidgetID == 'CallFunc_ClearColumnEndPoint':
            self.root.ids.DataBind_ColumnCellEndPoint.text = ''
        elif params_WidgetID == 'CallFunc_ClearRowStartPoint':
            self.root.ids.DataBind_RowCellStartPoint.text = ''
        elif params_WidgetID == 'CallFunc_ClearRowEndPoint':
            self.root.ids.DataBind_RowCellEndPoint.text = ''
        elif params_WidgetID == 'CallFunc_ExportProcess':
            pass
        elif params_WidgetID == 'CallFunc_ExportPathClear':
            self.root.ids.DataBind_FileExportPath.text = ''
            self.MDUserNotif_SnackbarHandler('Export Path Input has been cleared!')
        elif params_WidgetID == 'ClearInput_ExcelFileSheet':
            self.root.ids.DataBind_ExcelFileSheet.text = ''
            self.MDUserNotif_SnackbarHandler('Excel Sheet Name Input has been cleared!')
        elif params_WidgetID == 'CallFuncFromDataEditor_ClearInputs':
            self.root.ids.Resource_ItemNumVal.text = ''
            self.root.ids.Resource_ParticularProperty.text = ''
            self.root.ids.Resource_OnHandVal.text = ''
            self.root.ids.Resource_ProposedVal.text = ''
            self.root.ids.Resource_UnitTypeProperty.text = ''
            self.root.ids.Resource_UnitVal.text = ''
            self.root.ids.Total_ComputedVal.text = ''
            self.root.ids.Quartile_TextFieldVal_One.text = ''
            self.root.ids.Quartile_TextFieldVal_Two.text = ''
            self.root.ids.Quartile_TextFieldVal_Three.text = ''
            self.root.ids.Quartile_TextFieldVal_Four.text = ''
            self.MDUserNotif_SnackbarHandler('Text Field has been cleared!')
        else:
            raise ValueError('Parameter Variable -> params_TextFieldID: Received a string but no conditions were met.')
            KivyDebugger_Print.error('Parameter Variable -> params_TextFieldID: Received a string but no conditions were met.')
            exit()
            
    """
    (Future Functions) => <Various Functions>(self, ???), The Future Functions for the Future Version of this Program
        Why did you disable it? Because the client is only an individual hence having one of those functions at this point of time
        would just be a waste of time to produce content of the function body. There will be a chance to enable it for sometime but that depends
        on the Project Lead Decision whether he wants to continue this project or not. The function description is basically on the name itself.
    def MDAuthor_DeleteData(self):
        pass
    
    def MDAuthor_AddData(self):
        pass
    
    def MDAuthor_EditData(self):
        pass
    
    def MDStartup_FirstTime_AdminAccess(self, BooleanPrompt):
        pass

    def MDAuthor_CheckAdminPrev_OnAction(self):
        pass
    
    # For Logon Screen
    def MDAuthor_DialogBooleanPrompt(self):
        pass
    
    #For Logon and Admin for Modifying User and Assign
    def MDAuthor_PasswordAccessPrompt(self):
        pass
    For First TIme

    def MDNavigationDrawer_DynamicContent_Change(self, Params_Mode):
        if Params_Mode == 'User.Logon_Passed':
            #self.root.ids.
            pass
        elif Params_Mode == 'User.LogonPrompt':
            pass
        elif Params_Mode == 'User.PromptedAdmin_Passed':
            pass
        else:
            KivyDebugger_Print.error('RUNTIME ERROR: Unknown String Passed with No Valid Statement To Run!')
            raise ValueError('RUNTIME ERROR:  Unknown String Passed with No Valid Statement To Run!')
    
    """
"""
    NON-RELATABLE TO MAIN CLASS
        This are the class that has no connection with the main class that is used to initialize the whole program.

    NON-RELATABLE TO MAIN CLASS > AvatarSampleWidget Initializes an Item with Picture on the Left and Items to the right
    think of it as a MDListItem but with pictures on the left
"""
class AvatarSampleWidget(ILeftBody, Image):
    pass

"""
    NON-RELATABLE TO MAIN CLASS > IconLeftSampleWidget, Same Descriptions as AvatarSampleWidget, but it is only icons that gets displayed on the right
"""
class IconLeftSampleWidget(ILeftBodyTouch, MDIconButton):
    pass

"""
    NON-RELATABLE TO MAIN CLASS > IconRightSampleWidget, same as IconLeftSameplWidget
"""
class IconRightSampleWidget(IRightBodyTouch, MDCheckbox):
    pass

"""
    Non Class Functions, Unsorted, This is where things get started...
"""
if __name__ == '__main__':
    if MD_InventoryEXDI_SMLV_GUI().CriticalComponent_Check():
        KivyDebugger_Print.info('Component: All Set! Critical Component Passed.')
        MD_InventoryEXDI_SMLV_GUI().KivyConfig_Setup()
        MD_InventoryEXDI_SMLV_GUI().run()
    else:
        KivyDebugger_Print.error('CHECK FAILURE: Application failed failed to operate @ CriticalComponent_Check Function.')
        raise Exception('CHECK FAILURE: Application failed failed to operate @ CriticalComponent_Check Function.')
        exit(0)