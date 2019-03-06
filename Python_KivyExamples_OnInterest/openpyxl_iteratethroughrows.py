import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

ExcelFile = load_workbook('C:\\Users\\janja\\Desktop\\sam.xlsx', data_only = True)
ExcelWorksheet = ExcelFile['Sheet1']
#
#col_name = 'A'
#start_row = 1
#end_row = 99
#
#range_expr = "{col}{start_row}:{col}{end_row}".format(
#    col=col_name, start_row=start_row, end_row=end_row)
#
#for cell in ExcelWorksheet.iter_rows(range_string=range_expr):
#    print (cell.value)

for row_cells in ExcelWorksheet.iter_cols(min_col=column_index_from_string('A'), max_col=column_index_from_string('P'), min_row=6, max_row=25):
    for cell in row_cells:
       print('%s: cell.value=%s' % (cell, cell.value))